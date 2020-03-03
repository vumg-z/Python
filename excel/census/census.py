import openpyxl, os, pprint
from openpyxl.utils import get_column_letter, column_index_from_string


print(os.getcwd())

os.chdir(r"C:\Users\Billy\Desktop\B\Python/excel/census")

wb = openpyxl.load_workbook("censuspopdata.xlsx")

sheet = wb['Population by Census Tract']

countryData = {}

for row in range(2,sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    country = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    #make sure the key for state exists
    countryData.setdefault(state,{})
    # #make sure the key for this country exists
    countryData[state].setdefault(country, {'tracts': 0, 'pop': 0})
    #increase the country pop by the pop in this census tract.
    countryData[state][country]['tracts'] += 1
    #increase the country pop by the pop in the census track
    countryData[state][country]['pop'] += int(pop)

resultFile = open('census2010.py', 'w')
resultFile.write('allData=' + pprint.pformat(countryData))
resultFile.close()
